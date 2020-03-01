# -*- coding: utf-8 -*-
"""
Created on Mon Jun  3 22:01:50 2019
Spyderエディタ
For Python 3.7
@author: netchira
"""
#
# Input:
#    データ管理用のExcelファイル
#    上記Excelファイルで管理している項目の名前と、それが記入されている列番号を記載したテキストファイル
#
# 処理内容:
#    Excelを使用して何かを管理している業務を日常的に行っていると、
#    ある日時において管理していた情報と、1か月前に管理していた情報とに差分があることがある。
#    そのような場合に、Excelファイル同士の差分を抽出するのは困難であるため、一度、管理項目別に
#    sqlite3を使用してデータベース化し、データベースファイル同士を
#

class Excel2Database:
    # クラス変数を初期化
    KEYWORD_MAIN_CONTENTS = '[メイン項目]\n'
    KEYWORD_EXCEL_FILE_PATHS = '[DB化するExcelファイルパス一覧]\n'
    KEYWORD_SUB_CONTENTS = '[サブ項目]\n'
    KEYWORD_VAR_DB_START_INDEX = '[変数(DB登録開始行)]\n'
    
    def __init__(self):
        # インスタンス変数を初期化
        self.PARAM_DIC = {}
    
    def L_ReadText(self, txt_filepath):
        # txtファイルを読み込み、1行ずつ文字列型変数に格納し、
        # それらをtxtファイル全行分をリストに格納してリターンする。
        import os
        
        if os.path.exists(txt_filepath):            
            # ファイルをオープンする
            txt_data = open(txt_filepath, 'r')
            
            # 行ごとにすべて読み込んでリストデータにする
            txt_lines = txt_data.readlines()
            
            # 一行ずつ表示する
            #for line in txt_lines:
            #  print(line)
            
            # ファイルをクローズする
            txt_data.close()
        else:
            raise Warning('TXT File is not existing.')
        
        # 読み込んだデータをリターンする
        return txt_lines
    
    
    def L_SearchLineFeed(self, txt_file_content, start_idx): 
        # txtファイルに対して開始行を基準として、そこから初めて
        # 改行コードあるいはEOF(End Of File)が見つけられた行を
        # リターンする関数。ただし、本関数の引数にはリスト型で
        # テキストの内容を渡す必要がある。
        
        result_line_idx = 0
        
        if type(txt_file_content) == list:
            for line_idx in range(start_idx, len(txt_file_content)):
                # 最終行かどうか確認
                if line_idx == len(txt_file_content):
                    # 最終行であれば、最終行をリターンする。
                    result_line_idx = len(txt_file_content)
                elif (txt_file_content[line_idx] == '\n') or (txt_file_content[line_idx] == '\r\n'):
                    # 現在の行が改行コードであれば、その行をリターンする。
                    result_line_idx = line_idx
                    break
                else:
                    # 上記以外の場合、何もしない
                    pass
        else:
            # 与えられたテキストファイル内容がリスト型でなければ、
            # 例外を発生させる。
            raise Warning('L_SearchLineFeed Function Error !!')
    
        return result_line_idx
    
    
    def L_DetectExcelFilePaths(self, txt_lines):
        # txtファイルの一部を抽出し、Excelのファイルパスを格納したリストをリターンする。
        # ファイルパスに実際にファイルが存在するかどうか確認し、存在しなければ例外を出す。
        import os
        
        # txtファイルの一部を抽出するため、インデックスを取得
        xlsx_begin = txt_lines.index(self.KEYWORD_EXCEL_FILE_PATHS)
        xlsx_end = self.L_SearchLineFeed(txt_lines, xlsx_begin)
        
        # Excelファイルパスを取得していく
        xlsx_list = []
        for line_idx in range((xlsx_begin + 1), xlsx_end):
            line = txt_lines[line_idx]
            line = line.replace('\n','')
            if line != '':
                # ファイルパスを取得
                xl_filepath = os.path.abspath(line)
                if os.path.exists(xl_filepath):
                    # リストに追加
                    xlsx_list.append(line)
                else:
                    # 例外を発生させる
                    raise Warning('Excel File is not existing.\nError !!! -> ' + line)
            else:
                # 無視する
                pass
        
        
        return xlsx_list
        
    
    
    def L_DetectVarContents(self, txt_lines):    
        # txtファイルの一部を抽出
        var_begin = txt_lines.index(self.KEYWORD_VAR_DB_START_INDEX)
        line = txt_lines[(var_begin + 1)]
        line = line.replace(' ','')
        line = line.replace('行目','')
        line = line.replace('\n','')
        var_contents = line.split(':')
        return int(var_contents[1])
        
    
    def L_DetectMeetingTtileFromSubContents(self, txt_lines):
        # txtファイルの一部を抽出
        mtg_begin = txt_lines.index(self.KEYWORD_SUB_CONTENTS)
        line = txt_lines[(mtg_begin + 1)]
        line = line.replace(' ','')
        line = line.replace('行目','')
        line = line.replace('列目','')
        line = line.replace('\n','')
        mtg_contents = line.split(':')
        row, column = mtg_contents[1].split(',')
        return int(row), int(column)
        
            
    def L_DetectConfigFromSubContents(self, txt_lines, content_name):
        # 
        #
        #
        
        # content_name についてチェック
        if (content_name !='') and (content_name != None):
            # txtファイルの一部を抽出
            sub_begin = txt_lines.index(self.KEYWORD_SUB_CONTENTS)
            sub_end = self.L_SearchLineFeed(txt_lines, sub_begin)
            
            temp_dic = {}
            for line_idx in range((sub_begin + 1), sub_end):
                line = txt_lines[line_idx]
                line = line.replace(' ','')
                line = line.replace('行目','')
                line = line.replace('列目','')
                line = line.replace('\n','')
                contents = line.split(':')
                if contents != ['']:
                    # コロンで区切られた要素の内、1番目の要素をキーにして、2番目の要素をディクショナリ型変数に格納。
                    temp_dic[contents[0]] = contents[1]                                    
                else:
                    pass # temp_dic に要素を追加しない
        # content_nameがからっぽであった場合
        else:
            # 例外を発生させる。
            raise Warning('DetectConfigFromSubContents Function Error !!')
    
        return (temp_dic[content_name] == 'ON')
    
    
    def L_DetectMainContents(self, txt_lines):
        #
        # txtに記載している内容を解析し、インスタンス変数であるPARAM_DICを生成する。
        # このPARAM_DICには、以下の内容が格納されている。
        # 
        # key                    - value
        # Excelで管理したい項目の名前 - 「sqliteのコマンドで使用するデータのデータ型」と「Excelファイル内に記載されている列の数」 
        #
        # 約束事:
        # IDは自動的に付与する仕組みを実装するため、txtファイルには定義しなくてよい。
        # また、Excelに記載する項目の中で、txtファイルに規定した1番目がメイン項目として扱われ、
        # もしその項目がExcelファイル内で空欄となっていたら、その行のデータはすべてデータベースに取り込まないこととする。
        
        # txtファイルの一部を抽出
        main_begin = txt_lines.index(self.KEYWORD_MAIN_CONTENTS)
        main_end = self.L_SearchLineFeed(txt_lines, main_begin)
    
        self.PARAM_DIC = {}
        for line_idx in range((main_begin + 1), main_end):
            line = txt_lines[line_idx]
            line = line.replace(' ','')
            line = line.replace('列目','')
            line = line.replace('\n','')
            contents = line.split(':')
            if contents != ['']:
                # コロンで区切られた要素の内、3番目の要素をキーにして、4番目の要素(列)と2番目の要素(データ型)をセットにして
                #　ディクショナリ型変数に格納。
                self.PARAM_DIC[contents[2]] = (int(contents[3]), contents[1]) 
            else:
                pass # PARAM_DIC に要素を追加しない


    def L_MakeSQLStringForCreateTable(self):
        # 下記データを追加。
        # id:PRIMARY KEY指定, AUTO_INCREMENT指定
        # data_creation_date(データ生成日時):DEFAULT指定
        create_table_sql = '''CREATE TABLE main (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                          data_creation_date DEFAULT CURRENT_TIMESTAMP, '''
        # その他のデータはテキストファイルに記載されている情報をもとに追加。
        param_dic_list = list(self.PARAM_DIC.items())
        param_temp_list = []
        for val in param_dic_list:
            param_temp_list.append([val[1][0], val[0] + ' ' + val[1][1]])
        param_temp_list.sort()
        sql_str = ''
        for idx in range(len(param_temp_list)):
            sql_str = sql_str + param_temp_list[idx][1] + ', '
        sql_str = sql_str + 'meeting_title TEXT)'
        #todo TEXT, result TEXT, pic TEXT, deadline_date DATETIME, 
        #finished_date DATETIME, manager TEXT, checked_date DATETIME, 
        #note TEXT, meeting_title TEXT)
        create_table_sql = create_table_sql + sql_str
        # 最終結果をリターン
        return create_table_sql
        
    
    def L_CreataTables(self, sqlite_db, txt_lines):
        import sqlite3
        from contextlib import closing
        
        ok_flag = True
        try:
            # メイン項目を取得し、PARAM_DICを更新
            self.L_DetectMainContents(txt_lines)
            
            with closing(sqlite3.connect(sqlite_db)) as conn:
                c = conn.cursor()
                # executeメソッドでSQL文を実行する
                # [CREATE TABLE]
                create_table_sql = self.L_MakeSQLStringForCreateTable()
                c.execute(create_table_sql)
            
                # executeメソッドでSQL文を実行する
                # [CREATE INDEX]
                # INDEX指定:id(AUTO_INCREMENT指定をしたいため、しぶしぶ設定)
                create_index = 'CREATE INDEX id_index ON main(id)'
                c.execute(create_index)
            
                # executeメソッドでSQL文を実行する
                # [CREATE TABLE]
                # id:PRIMARY KEY指定, AUTO_INCREMENT指定
                create_table = '''CREATE TABLE sub (meeting_id INTEGER PRIMARY KEY AUTOINCREMENT, 
                               data_creation_date DEFAULT CURRENT_TIMESTAMP, 
                               meeting_title TEXT, excel_file_name TEXT,
                               excel_update_date DATETIME)'''
                c.execute(create_table)
            
                # executeメソッドでSQL文を実行する
                # [CREATE INDEX]
                # INDEX指定:meeting_id_index(AUTO_INCREMENT指定をしたいため、しぶしぶ設定)
                create_index = 'CREATE INDEX meeting_id_index ON sub(meeting_id)'
                c.execute(create_index)
                    
                # Save (commit) the changes, and close.
                conn.commit()
                conn.close()
                
        except:
            ok_flag = False
        
        return ok_flag
            
    
    
    def L_MakeSQLStringForInsertInto(self):
        # SQL文に値をセットする場合は，Pythonのformatメソッドなどは使わずに，
        # セットしたい場所に?を記述し，executeメソッドの第2引数に?に当てはめる値を
        # タプルで渡す.(INSERT文におけるSQLインジェクション脆弱性への対策として)

        # SQLコマンド文の見本
        # INSERT INTO main ( todo, result, pic, deadline_date, finished_date, manager, 
        #    checked_date, note, meeting_title)  VALUES (?,?,?,?,?,?,?,?,?)

        # PARAM_DICから、[ (dbに登録する)    ]
        param_dic_list = list(self.PARAM_DIC.items())
        param_temp_list = []
        for val in param_dic_list:
            param_temp_list.append([val[1][0], val[0]])
        param_temp_list.sort()
        
        # [INSERT INTO]を実行する際の変数名の指定をする分の生成
        insert_sql = 'INSERT INTO main ('
        sql_str = ''
        for idx in range(len(param_temp_list)):
            sql_str = sql_str + param_temp_list[idx][1] + ', '
        sql_str = sql_str + 'meeting_title)'
        
        insert_sql = insert_sql + sql_str
        insert_sql = insert_sql + ' VALUES ('
        question_str = '?,' * len(param_temp_list)
        question_str = question_str + '?)'
        insert_sql = insert_sql + question_str
        
        # 最終結果をリターン
        return insert_sql
    
    
    def L_GenerateInsertionData(self, xl_sheet_obj, first_row, meeting_title):
        #
        # PARAM_DICを使用してExcelファイルを解析し、データベースへ登録するためのでデータを準備する。
        # main TABLEの構成要素の「列数」を用いてExcelを読みに行き、セルのデータの取得する。
        #
        
        param_val_list = list(self.PARAM_DIC.values())
        param_col_list = []
        for val in param_val_list:
            param_col = val[0] # Excelで管理している項目の「列数」を取得
            param_col_list.append(param_col)
    
        insert_dara_list = []
        for row_idx in range(first_row, (xl_sheet_obj.max_row + 1)):
            temp_list = []
            for col_idx in param_col_list:
                # Excelのセルを1列ずつ順番(昇順)に読み込み、値をリストに追加
                temp_list.append(xl_sheet_obj.cell(row = row_idx, column = col_idx).value)
            temp_list.append(meeting_title) # 最後に会議名を追加
            temp_tup = tuple(temp_list)
            if temp_tup[0] != None: # txtファイルに規定した1番目の要素が空欄であるかどうか確認
                insert_dara_list.append(temp_tup)
            else:
                pass # txtファイルに規定した1番目の要素がNoneだった場合、それはInsertDataとして準備しないこととする。
        # 最終結果をリターン
        return insert_dara_list
    
    
    def L_InsertDataToTables(self, db, txt_lines):
        import os
        import datetime
        import sqlite3
        from contextlib import closing
        import openpyxl as px
    
        # Excelファイル一覧
        xlsx_list = self.L_DetectExcelFilePaths(txt_lines)
        
        # 変数
        first_row = self.L_DetectVarContents(txt_lines)
        #first_row = 8 # 8行目を表す。1オリジン。Excelベースの行数で良い。
        
        # メイン項目
        self.L_DetectMainContents(txt_lines)
        
        for xlsx in xlsx_list:
            # Excelファイルを開き、シート情報を取得する
            book = px.load_workbook(xlsx)
            sheets = book.sheetnames
            first_sheet = book[sheets[0]] # 1番目のシートを取得
            
            # サブ項目 1
            mtg_row, mtg_col = self.L_DetectMeetingTtileFromSubContents(txt_lines)
            meeting_title = first_sheet.cell(row = mtg_row, column = mtg_col).value
            # サブ項目 2
            if self.L_DetectConfigFromSubContents(txt_lines, u'Excelファイル名'):
                xl_filename = os.path.basename(xlsx)
            # サブ項目 3
            if self.L_DetectConfigFromSubContents(txt_lines, u'Excelファイル最終更新日時'):
                xl_file_datetime = datetime.datetime.fromtimestamp(os.stat(xlsx).st_mtime)
                dt = xl_file_datetime - datetime.timedelta(hours=9) # db内で使用される時刻は、ＵＴＣで統一する。 
                xl_update_date = dt.strftime('%Y-%m-%d %H:%M:%S')
            
            # ベースDBに対しデータを追加する
            with closing(sqlite3.connect(db)) as conn:
                c = conn.cursor()
        
                # 一度に複数のSQL文を実行したいときは，タプルのリストを作成した上で
                # executemanyメソッドを実行する.
                # [INSERT INTO]
                insert_sql = self.L_MakeSQLStringForInsertInto()
                data = self.L_GenerateInsertionData(first_sheet, first_row, meeting_title)
                c.executemany(insert_sql, data)
                    
                # executeメソッドでSQL文を実行する
                # [INSERT INTO]
                insert_sql = '''INSERT INTO sub (meeting_title, excel_file_name, 
                                excel_update_date) VALUES (?,?,?)'''
                data = (meeting_title, xl_filename, xl_update_date)        
                c.execute(insert_sql, data)
            
                # Save (commit) the changes, and close.
                conn.commit()
                conn.close()
    
    
    def ExecExcel2Database(self, txt_filepath):
        # 注意事項
        # db内で使用される時刻は、UTCで統一する。 
        #
        #
        # https://docs.microsoft.com/ja-jp/sql/t-sql/data-types/datetime-transact-sql?view=sql-server-2017
        # SQL datetimeが表せる文字列リテラル　=> YYYYMMDD hh:mm:ss[.mmm]
        
        import datetime
        import os
        
        # テキストファイルの内容を取得
        txt_lines = self.L_ReadText(txt_filepath)
    
        # 現在時刻からTemp用のdbを生成
        today = datetime.datetime.today()
        today_str = '{:%Y%m%d%H%M%S}'.format(today)
        tempdb_str = r'C:\PyTemp\Temp_' + today_str + '.db'
        tempdb_filepath = os.path.abspath(tempdb_str)
        
        # ベースとなるDBを生成
        ok_flag = self.L_CreataTables(tempdb_filepath, txt_lines)
        
        if ok_flag:
            # DBにデータを追加
            self.L_InsertDataToTables(tempdb_filepath, txt_lines)
        else:
            pass
        
        print('Creat DB Finished !!!')
        # 作成したdbファイルのパスをリターンする。
        return tempdb_filepath


    def ShowTables(self, db, select_condition):
        #
        # dbの内容を表示させる(SELECTコマンドを実行する)
        #
        
        import sqlite3
        from contextlib import closing
    
        # DBの内容を表示
        with closing(sqlite3.connect(db)) as conn:
            c = conn.cursor()
        
            # executeメソッドでSQL文を実行する
            # [SELECT * FROM]
            select_sql = 'SELECT * FROM  sqlite_master' # WHERE type=''TABLE'' and name=''main'';
            for row in c.execute(select_sql):
                print(row)
            
            # executeメソッドでSQL文を実行する
            # [SELECT * FROM]
            print('----------- SELECT WHERE (main)-----------')           
            # select_condition がNoneであれば単純なSELECTを実行する
            if select_condition is None:
                select_sql = 'SELECT * FROM main'
            # その他の場合は、select_conditionを用いてSELECT WHEREを実行する
            else:
                select_sql = 'SELECT * FROM main WHERE ' + select_condition
            for row in c.execute(select_sql):
                print(row)
            print('----------- SELECT WHERE (main)-----------')
            
            print('----------- SELECT (sub)-----------')           
            # executeメソッドでSQL文を実行する
            # [SELECT * FROM]
            select_sql = 'SELECT * FROM sub'
            for row in c.execute(select_sql):
                print(row)
            print('----------- SELECT (sub)-----------')           
    
            print('----------- INNER JOIN -----------')
            # executeメソッドでSQL文を実行する
            # [SELECT * FROM]
            select_sql = 'SELECT * FROM main INNER JOIN sub ON main.meeting_title = sub.meeting_title'
            for row in c.execute(select_sql):
                print(row)
            print('----------- INNER JOIN -----------')
            
            # Save (commit) the changes, and close.
            conn.commit()
            conn.close()
    

    def ExecSqlDiff(self, db1, db2):
        #
        # 引数に与えた2つのデータベースファイルに対し、sqlite3インストール時に付属してくるsqldiff.exeを
        # 使用して差分を表示させる。
        #
        import subprocess
        import os
        sqldiff_exe_path = r'C:\Program Files (x86)\sqlite-tools\sqldiff.exe'
        if (os.path.exists(db1)) and (os.path.exists(db2)):
            cmd = sqldiff_exe_path + ' ' +  db1 + ' ' + db2
            #cmd = [sqldiff_exe_path, db1, db2] # これでも動作する
            returncode = subprocess.Popen(cmd, stdout=subprocess.PIPE)
            print("###################################################################")
            for line in returncode.stdout:
                print("#" + line.decode(encoding='utf-8'))
            print("###################################################################")

        else:
            print('.db file is not found.')


# main
txt_filepath = r'C:\PyTemp\ExcelFormat.txt'
a = Excel2Database()
print(a.PARAM_DIC)
tempdb_filepath = a.ExecExcel2Database(txt_filepath)

# DBの内容を表示
a.ShowTables(tempdb_filepath, None)

# DBの内容を表示
a.ShowTables(tempdb_filepath, 'deadline_date < CURRENT_DATE')
a.ShowTables(tempdb_filepath, 'deadline_date < datetime(''2020-01-05'');')

db1 = r'C:\PyTemp\Temp_20190615132454.db'
db2 = r'C:\PyTemp\Temp_20200301130813.db'
a.ExecSqlDiff(db1, db2)






