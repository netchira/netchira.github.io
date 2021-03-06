﻿# -*- coding: utf-8 -*-
"""
Created on Fri Dec 29 20:28:35 2017
Spyderエディタ
For Python 2.7
@author: netchira
"""
import os
import Tkinter      #GUIコントロール
import ttk          #GUIコントロール(Combobox)
import tkFileDialog #ファイル操作
import tkMessageBox #メッセージボックス表示
import openpyxl as px
from openpyxl.drawing.image import Image
import PIL.Image
import numpy as np
import re
from glob import glob

#スクリプトの開始
print "Insert Pictures In Excel Start"

#root定義
root = Tkinter.Tk()
root.title(u"画像ファイル自動挿入スクリプト")
root.geometry("600x300")

#定数の定義
DEFAULT_SIZE = '500'

#色の定義
BLACK = '#000000'
RED = '#ff0000'
PINK = '#ffaacc'

#root上のグローバル変数定義
g_PictureType =Tkinter.StringVar()
g_PictureType.set('type')#変数の初期化
g_ExcelFile =Tkinter.StringVar()
g_ExcelFile.set('file_name(output)')#変数の初期化
g_PictFolderPath =Tkinter.StringVar()
g_PictFolderPath.set('folder_name(input)')#変数の初期化
g_SelectedSheet =Tkinter.StringVar()
g_SelectedSheet.set('nothing')#変数の初期化

g_PictSizeWidth =Tkinter.StringVar()
g_PictSizeWidth.set(DEFAULT_SIZE)#変数の初期化
g_PictSizeHeight =Tkinter.StringVar()
g_PictSizeHeight.set(DEFAULT_SIZE)#変数の初期化
g_PictPosition =Tkinter.StringVar()
g_PictPosition.set('A10,L11')#変数の初期化


g_InsertionPicturesName =Tkinter.StringVar()

g_ExcelSheets = ('A', 'B', 'C') #tuple 初期化

"""
 呼び出し元：ボタン1
 関数の処理内容：Excelファイルを選択する
"""
def SelectExcelFile(event):
    global g_ExcelFile

    #Entryに文字列がタイプされているか確認
    if (Entry1.get() != '') & os.path.exists(Entry1.get()):
        iDir = os.path.abspath(Entry1.get())
        iFile = (Entry1.get()).split('/')[-1]
    else:
        iDir = os.path.abspath(os.path.curdir)
        iFile = ''
    #tkMessageBox.showinfo('Excelファイル選択','ファイルを選択してください。')
    fTyp = [("xlsxファイル",".xlsx"),("xlsファイル",".xls"),("xlsmファイル",".xlsm"),("すべてのファイル",".*")]
    
    #Excelファイルを選択する
    filepath = tkFileDialog.askopenfilename(filetypes = fTyp, initialdir = iDir, initialfile = iFile)
    if filepath != "":
        g_ExcelFile.set(filepath)
        Entry1.configure(textvariable = g_ExcelFile)
    
    #Excelファイルからシート名を抽出
    if os.path.exists(Entry1.get()):
        wb = px.load_workbook(Entry1.get())
        sheet_list = wb.get_sheet_names()
        sheet_tuple = tuple(sheet_list)
        g_ExcelSheets = sheet_tuple
        print g_ExcelSheets
    else:
        g_ExcelSheets = ('A', 'B', 'C')
    
    # Combobox の選択項目を更新
    Combobox2['values'] = g_ExcelSheets
    Combobox2.set(g_ExcelSheets[0])

"""
 呼び出し元：ボタン2
 関数の処理内容：画像ファイルが格納されているフォルダを選択する
"""
def SelectFolder(event):
    global g_PictFolderPath
    #フォルダを選択する
    if (Entry2.get() != '') & os.path.exists(Entry2.get()):
        iDir = os.path.abspath(Entry2.get())
    else:
        iDir = os.path.abspath(os.path.curdir)
    #tkMessageBox.showinfo('フォルダ選択','画像ファイルが格納されているフォルダを選択してください。')
    folderpath = tkFileDialog.askdirectory(initialdir = iDir)
    if folderpath != "":
        g_PictFolderPath.set(folderpath)
        Entry2.configure(textvariable = g_PictFolderPath)

    # 指定したフォルダ内にある画像ファイル一覧を取得する
    picts_list = get_picture_list()
    if picts_list != None:   
        # リストボックス変数へのセット
        g_InsertionPicturesName.set(tuple(picts_list))

    
"""
 呼び出し元：ボタン3
 関数の処理内容：Excelファイルに画像を挿入する
"""
def InsertPictureFiles(event):
    global g_PictPosition
    #global g_InsertionPicturesName
    
    l_pic_position = g_PictPosition.get()
    #l_picts_list = g_InsertionPicturesName.get() # 1文字ずつ取得してしまい、誤実装
    
    # 画像ファイル一覧を取得
    picts_list = get_picture_list() 
    
    # ワークブックwb,ワークシートwsを取得
    xl_list = get_excel_sheet()

    # 変数が空でないことを確認
    if (picts_list != None) & (xl_list != None):
        # xl_listを分解
        wb = xl_list[0]
        ws = xl_list[1]
    
        # 画像の挿入位置の算出
        picts_num = len(picts_list)
        position_list = CalculatePosition(l_pic_position, picts_num)

        # サイズ編集後の画像ファイル格納フォルダパスを取得
        new_directory = make_new_directory()

        # Excelファイルの編集処理
        for index, pict in enumerate(picts_list):
            # 画像の編集
            PictureSizeConversion(pict, new_directory)
            
            # 画像ファイル名の挿入
            filename_pos = position_list[index]
            ws[filename_pos] = GetFileName(pict)
            
            # 画像の挿入
            image_pos = CellRowIncrement(position_list[index])
            new_pic_path = new_directory +  "/" + os.path.basename(pict)
            temp_img = Image(new_pic_path)
            ws.add_image(temp_img, image_pos)
            
        # Excelファイルを新規作成・保存
        newwb = save_excel_file(wb)
        
        # メッセージ表示
        print "Insert Pictures Success!"
        dispmsg = u"Excelファイルに画像を挿入しました。ベースとなったExcelファイルと同じディレクトリに新しいExcelファイルを出力しました。\n"
        dispmsg += u"また、一時的にサイズを変更した画像ファイルをTempForPythonフォルダに格納しています。不要であれば削除してください。\n"
        dispmsg += u"\n"
        dispmsg += u"出力ファイル名:\n"
        dispmsg += ( newwb + "\n\n")
        dispmsg += u"サイズ編集後の画像ファイル出力先:\n"        
        dispmsg += ( new_directory + "\n")
        tkMessageBox.showinfo('Success', dispmsg)
    
    else:
        # メッセージ表示
        dispmsg = u"Excelファイルまたはフォルダが適切ではありませんでした。処理を中止します。\n"
        tkMessageBox.showinfo('Failure', dispmsg)

"""
 関数の処理内容：指定したフォルダが存在しているか確認し、そのフォルダ内にある画像ファイルの一覧をリターンする。
"""
def get_picture_list():
    global g_PictFolderPath
    global g_PictureType

    l_folder = g_PictFolderPath.get()
    l_type   = g_PictureType.get()
    
    #フォルダの所在確認
    l_folder_exist_flag = os.path.isdir(l_folder)
   
    if l_folder_exist_flag:               
        # 画像ファイルのみのリストを作成する(拡張子はユーザーが指定したもの)
        picts_path = l_folder + '/*'
        picts_path += l_type
        picts_list = glob(picts_path)
        
    else:
        # エラー表示：存在していないフォルダ名の指定
        dispmsg = u"選択したフォルダが存在していませんでした。再度、フォルダを選択してください。\n"
        dispmsg += (l_folder + "\n")
        tkMessageBox.showinfo(u'画像を格納したフォルダが見つかりません', dispmsg)
        picts_list = []
        
    return picts_list
        
"""
 関数の処理内容：指定したフォルダが存在しているか確認し、そのフォルダ内に新しいディレクトリを作成し、そのフォルダパスをリターンする。
"""
def make_new_directory():
    global g_PictFolderPath

    l_folder = g_PictFolderPath.get()
    
    #フォルダの所在確認
    l_folder_exist_flag = os.path.isdir(l_folder)
   
    if l_folder_exist_flag:       
        # 画像の一時保存用フォルダを作成する
        new_directory = l_folder + u"/TempForPython"
        if not os.path.exists(new_directory):
            os.mkdir(new_directory)
            
    else:
        # エラー表示：存在していないフォルダ名の指定
        dispmsg = u"選択したフォルダが存在していませんでした。再度、フォルダを選択してください。\n"
        dispmsg += (l_folder + "\n")
        tkMessageBox.showinfo(u'画像を格納したフォルダが見つかりません', dispmsg)
        new_directory = []
        
    return new_directory
        

"""
 関数の処理内容：Excelファイルを指定し、ワークシートをリターンする
"""
def get_excel_sheet():
    global g_ExcelFile
    global g_SelectedSheet
    
    l_file   = g_ExcelFile.get()
    l_selectedsheet = g_SelectedSheet.get()
    
    ret_list = []
    
    #ファイルの所在確認
    l_excel_exist_flag = os.path.isfile(l_file)
   
    if l_excel_exist_flag:               
        # Excelファイルを開く
        wb = px.load_workbook(l_file)

        # シートを指定する
        ws = wb.get_sheet_by_name(l_selectedsheet)
        
        ret_list.append(wb)
        ret_list.append(ws)
    else:
        # エラー表示：存在していないフォルダ名の指定
        dispmsg = u"選択したExcelファイルが存在していませんでした。再度、ファイルを選択してください。\n"
        dispmsg += (l_file + "\n")
        tkMessageBox.showinfo(u'指定したExcelファイルが見つかりません', dispmsg)
        ret_list = []

    return ret_list

"""
 関数の処理内容：Excelファイルを保存し、新しいExcelファイルのパスをリターンする
"""
def save_excel_file(wb):
    global g_ExcelFile
    global g_PictFolderPath
    
    l_file   = g_ExcelFile.get()
    l_folder = g_PictFolderPath.get()
    
    # 新しいExcelファイルを保存(ファイル名の末尾に'_new'を追加)
    temp_filename = GetFileName(l_file)
    temp_filename = l_folder + "/" + temp_filename
    temp_file_extention = GetFileExtention(l_file)
    newwb = temp_filename  + "_new." + temp_file_extention
    wb.save(newwb)
    
    return newwb


"""
 関数の処理内容：英字だけを抽出し、リストで返す。([a-zA-Z]は正規表現)
"""
def get_alphabet_list(text):
    matchedList = re.findall('[a-zA-Z]',text)
    return matchedList

"""
 関数の処理内容：数字だけを抽出し、リストで返す。([0-9]は正規表現)
"""
def get_number_list(text):
    temp = re.findall('[0-9]*',text)
    numberList = filter(None, temp)
    return numberList

"""
 関数の処理内容：PILライブラリを使って画像のサイズを変更し、新しいファイルとして保存する
"""
def PictureSizeConversion(pic_path, hozon_directory):
    global g_PictSizeHeight
    global g_PictSizeWidth
    
    # 変更したい画像サイズの取得
    if Entry3.get() == '':
        g_PictSizeHeight.set(DEFAULT_SIZE)
        l_pic_height = g_PictSizeHeight.get()
    else:
        g_PictSizeHeight.set(Entry3.get())
        l_pic_height = g_PictSizeHeight.get()
        
    if Entry4.get() == '':
        g_PictSizeWidth.set(DEFAULT_SIZE)
        l_pic_width  = g_PictSizeWidth.get()
    else:
        g_PictSizeWidth.set(Entry4.get())
        l_pic_width = g_PictSizeWidth.get()
        
    l_intHeight = int(l_pic_height)
    l_intWidth  = int(l_pic_width)
    
    if (l_intHeight > 0) & (l_intWidth > 0):
        # 画像のサイズ変更(PILライブラリ)
        origin_pic_path = pic_path
        img_raw = PIL.Image.open(origin_pic_path)
        img_copy = img_raw.copy()
        img_resize = img_copy.resize((l_intWidth, l_intHeight), PIL.Image.LANCZOS)
        
        # 画像の保存
        new_pic_path = hozon_directory + "/" + os.path.basename(pic_path)
        img_resize.save(new_pic_path)
    else:
        # エラー表示：画像サイズの不適切な指定
        dispmsg = u"画像のサイズ(縦・横)がゼロとなっています。再度指定しなおしてください。\n"
        tkMessageBox.showinfo(u'画像サイズの不適切な指定', dispmsg)

"""
 関数の処理内容：画像を挿入するセルの位置を算出する
"""
def CalculatePosition(base_position, picts_num):   
    global g_PictSizeHeight
    
    # 画像のサイズからExcel何行分に相当するか算出
    
    if Entry3.get() == '':
        g_PictSizeHeight.set(DEFAULT_SIZE)
        l_pic_height = g_PictSizeHeight.get()
    else:
        g_PictSizeHeight.set(Entry3.get())
        l_pic_height = g_PictSizeHeight.get()
        
    l_intHeight = int(l_pic_height)
    l_pict_height = l_intHeight / 16
    
    # ユーザーが指定した挿入位置の決定するために、行列を作成して何行何列か算出する
    split_str_list = base_position.split(',')
    column = len(split_str_list)
    
    if picts_num % column == 0:
        row = picts_num / column
    else:
        row = picts_num / column + 1     
    
    # ユーザーが指定したセル位置の列数のみ(数字部分のみ)を取得し、NumPy行列の生成
    number_list = get_number_list(base_position)
    nparray = np.array(number_list, dtype = 'int64')
    
    # NumPy行列を必要列数分リピートする。
    xl_base_array = np.tile(nparray, row)
    
    # 画像を張り付ける2行目以降の位置を決定
    pict_offset_postion_array = np.arange(0, l_pict_height * row, l_pict_height)
    xl_cell_number_array = pict_offset_postion_array.repeat(column)
    
    # 行列の加算、リストへの変換
    xl_target_cell_array = xl_base_array + xl_cell_number_array 
    xl_target_cell_list = xl_target_cell_array.tolist()
    
    # str型に変換
    result_sono1= []
    for content in xl_target_cell_list:
        result_sono1.append(str(content))
        
    # 半角英字だけ抽出
    text_list = get_alphabet_list(base_position)
    result_sono2= []
    for x in range(0, row):
        for index in range(0, column):
            result_sono2.append(text_list[index])
    
    # 最終的な結果を算出
    result = []
    if len(result_sono1) == len(result_sono2):
        for index in range(0, len(result_sono1)):
            result.append(result_sono2[index] + result_sono1[index])
    else:
        result = None
    
    # リストの要素数を、画像の個数分だけに絞る
    result = result[:picts_num]
    
    print result
    return result

"""
 関数の処理内容：セルの位置(行のみ)をインクリメントする
"""
def CellRowIncrement(cell):
    collumn = get_alphabet_list(cell)
    row_old = get_number_list(cell)
    
    # 行の値だけインクリメントし、セル位置を新しく文字列で生成
    incremented = int(row_old[0]) + 1
    row_new = str(incremented)
    cell_new = collumn[0] + row_new
    return cell_new

"""
関数の処理内容：ファイルのフルパスから拡張子を除いたファイル名だけを返す
"""
def GetFileName(file_path):
    file_name_ext = os.path.basename(file_path)
    temp = file_name_ext.split(".")
    file_name = temp[0] # 拡張子のないファイル名を取得
    return file_name

"""
関数の処理内容：ファイルのフルパスから拡張子を除いたファイル名だけを返す
"""
def GetFileExtention(file_path):
    file_name_ext = os.path.basename(file_path)
    temp = file_name_ext.split(".")
    file_extention = temp[-1] # ファイルの拡張子を取得
    return file_extention


#ラベル1(textblock)を配置
dispmsg = u"Excelファイルと、画像ファイルが格納されたフォルダを選択してください。その後、画像挿入をクリックしてください。"
Label1 = Tkinter.Label(text=dispmsg, foreground=BLACK, background=PINK)
Label1.place(x=0, y=0)

#ラベル2(textblock)を配置
Label2 = Tkinter.Label(text=u'拡張子：', width=10, anchor=Tkinter.W)
Label2.place(x=30, y=30)

#ラベル3(textblock)を配置
Label3 = Tkinter.Label(text=u'サイズ縦：', width=10, anchor=Tkinter.W)
Label3.place(x=210, y=30)

#ラベル4(textblock)を配置
Label4 = Tkinter.Label(text=u'サイズ横：', width=10, anchor=Tkinter.W)
Label4.place(x=370, y=30)

#ラベル5(textblock)を配置
Label5 = Tkinter.Label(text=u'シート名：', width=10, anchor=Tkinter.W)
Label5.place(x=30, y=90)

#ラベル6(textblock)を配置
Label6 = Tkinter.Label(text=u'挿入セル：', width=10, anchor=Tkinter.W)
Label6.place(x=210, y=90)

#ラベル7(textblock)を配置
Label7 = Tkinter.Label(text=u'(※コンマ(,)で区切ると複数列に挿入)')
Label7.place(x=340, y=90)

#ラベル8(textblock)を配置
Label8 = Tkinter.Label(text=u'画像ファイル：')
Label8.place(x=10, y=150)

#コンボボックス1(combobox)を配置
Combobox1= ttk.Combobox(width=10, textvariable=g_PictureType)
Combobox1.bind('<<ComboboxSelected>>')
Combobox1['values']=('.png','.jpeg','.bmp')
Combobox1.set('.png')
Combobox1.place(x=90, y=30)

#コンボボックス2(combobox)を配置
Combobox2= ttk.Combobox(width=10, textvariable=g_SelectedSheet)
Combobox2.bind('<<ComboboxSelected>>')
Combobox2['values']=g_ExcelSheets
Combobox2.set('')
Combobox2.place(x=90, y=90)

#ボタン1(button)を配置
Button1 = Tkinter.Button(width=10, text=u'ファイル選択')
Button1.bind("<ButtonRelease-1>",SelectExcelFile) #左クリックされると関数を呼び出すようにバインド
Button1.place(x=0, y=60)
    
#ボタン2(button)を配置
Button2 = Tkinter.Button(width=10, text=u'フォルダ指定')
Button2.bind("<ButtonRelease-1>",SelectFolder) #左クリックされると関数を呼び出すようにバインド
Button2.place(x=0, y=120)

#ボタン3(button)を配置
Button3 = Tkinter.Button(width=10, text=u'画像挿入')
Button3.bind("<ButtonRelease-1>",InsertPictureFiles) #左クリックされると関数を呼び出すようにバインド
Button3.place(x=0, y=270)

# リストボックス用にフレーム(Frame)を配置
Frame = Tkinter.Frame()
Frame.place(x=90, y=150)

# スクロールバー付きリストボックス(Listbox)を配置
Listbox = Tkinter.Listbox(Frame, width = 80, height=7, listvariable=g_InsertionPicturesName)
Listbox.grid(row=0, column=0)

# スクロールバー(Scrollbar)を配置
Scrollbar = ttk.Scrollbar(Frame, orient=Tkinter.VERTICAL, command=Listbox.yview)
Scrollbar.grid(row=0, column=1, sticky=(Tkinter.N + Tkinter.S))
Listbox['yscrollcommand'] = Scrollbar.set

#エントリー1(textbox)を配置
Entry1 = Tkinter.Entry(width=80, textvariable=g_ExcelFile)
Entry1.place(x=90, y=60)

#エントリー2(textbox)を配置
Entry2 = Tkinter.Entry(width=80, textvariable=g_PictFolderPath)
Entry2.place(x=90, y=120)

#エントリー3(textbox)を配置
Entry3 = Tkinter.Entry(width=10, textvariable=g_PictSizeHeight)
Entry3.place(x=270, y=30)

#エントリー4(textbox)を配置
Entry4 = Tkinter.Entry(width=10, textvariable=g_PictSizeWidth)
Entry4.place(x=430, y=30)

#エントリー5(textbox)を配置
Entry5 = Tkinter.Entry(width=10, textvariable=g_PictPosition)
Entry5.place(x=270, y=90)


#mainloop
root.mainloop()